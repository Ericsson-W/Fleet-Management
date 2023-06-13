# Imports
from os import close
from sys import path
import numpy as np
import time
import math
import datetime

import matplotlib.pyplot as plt
import utils.pre_processing as dyn
from robots.duckietown import duckiebot_inverse_kinematics
from robots.PID import PID
from robots.RRT import RRT
import openpyxl
from openpyxl import load_workbook


# TODO: make something to deal with ROS custom message setup for computers (script?)

###########################################
######## Individual Controllers ###########
###########################################

# In run_live.py these are called once per self.controlled_vehicles[vehicle].

                        #     "v_bar": 0.5, "distance2front": 0.25,
                        #     "lookahead_distance": 0.25, "intersection_timeoff": 2.5,
                        #     "path_buffer": 0.5,
                        #     "PID": {"p_gain": +6.9, "i_gain": +1.1, "d_gain": +1.3,
                        #         "windup_measures": True, "error_sum_max": 1.0}
                        # }
class constant_linear_vel():
    """ Outputs a constant linear velocity.
        For debug purposes only. """
    
    def __init__(self, controller_parameters: dict, path_parameters: dict, path_generator, controlled_vehicles, goal_paths, delta_t):
        
        self.v_bar:float        = controller_parameters["v_bar"]
        
        self.buffer_points      = int(path_parameters['path_points_per_m'] * controller_parameters['path_buffer'])
        self.lookahead_points   = int(path_parameters['path_points_per_m'] * controller_parameters['lookahead_distance'])

        # Take time step to be constant during experiment
        self.delta_t = delta_t

        self.goal_paths             = goal_paths


        self.vehicles_control_info = {
            vehicle: {"closest_path_point_index": None, "target_pos": None,
            "lateral_controller": PID(controller_parameters['PID'])} 
            for vehicle in self.controlled_vehicles.keys()
        }
        
    def control_action(self, vehicle):
        """ Uses info/values obtained from the control_step() to calculate the
            desired linear and angular velocities for the vehicles.
        """

        # Base value for v
        v = self.v_bar

        # TODO: why does self.controlled_vehicles[vehicle].position not work? need self.controlled_vechiles[vehicle].position?
        angle2target = dyn.angle(self.vehicles_control_info[vehicle]['target_pos'] - self.controlled_vehicles[vehicle].position, radians=True)
        heading_error = angle_within_range(angle2target - self.controlled_vehicles[vehicle].angle) # Range [-180, +180]
        
        # Base value for omega
        omega = self.vehicles_control_info[vehicle]['lateral_controller'].control_step(heading_error, self.delta_t) # PID controller

        return v, omega


    def control_step(self):

            for vehicle in self.controlled_vehicles:
                            
                
                # Selecting sub-section of path where the goal point can be located.
                ## Because of path overlap at intersection, basic min distance would not work.
                if self.vehicles_control_info[vehicle]['closest_path_point_index'] is None:
                    goal_path_subset = self.goal_paths[vehicle]

                else:
                    path_buffer_end_idx = self.vehicles_control_info[vehicle]['closest_path_point_index'] + self.buffer_points

                    # TODO: add something for when path_loop = False
                    if path_buffer_end_idx < len(self.goal_paths[vehicle]):
                        goal_path_subset = self.goal_paths[vehicle][self.vehicles_control_info[vehicle]['closest_path_point_index']:path_buffer_end_idx]
                    else:
                        overflow = path_buffer_end_idx - len(self.goal_paths[vehicle])
                        goal_path_subset = self.goal_paths[vehicle][self.vehicles_control_info[vehicle]['closest_path_point_index']:len(self.goal_paths[vehicle])]
                        goal_path_subset = np.concatenate([goal_path_subset, self.goal_paths[vehicle][:overflow]])
                ###############

                # Getting index of closest point in sub-path
                distance_to_subpath_points = np.linalg.norm(self.controlled_vehicles[vehicle].position - goal_path_subset, axis=1)
                min_distance = np.min(distance_to_subpath_points)
                closest_pathpoint_idx = np.where(distance_to_subpath_points == min_distance)[0][0]

                # Getting index of closest point in goal path
                if self.vehicles_control_info[vehicle]['closest_path_point_index'] is not None:
                    closest_pathpoint_idx = (closest_pathpoint_idx + self.vehicles_control_info[vehicle]['closest_path_point_index']) % (len(self.goal_paths[vehicle])-1)

                self.vehicles_control_info[vehicle]['closest_path_point_index'] = closest_pathpoint_idx


                # Getting target position
                target_idx = (self.vehicles_control_info[vehicle]['closest_path_point_index'] + self.lookahead_points) % (len(self.goal_paths[vehicle]) - 1)
                self.vehicles_control_info[vehicle]['target_pos'] = self.goal_paths[vehicle][target_idx]

                # Using info obtained above to get control action
                v, omega = self.control_action(vehicle)

                # Robot object converts (v,omega) to the required ROS command
                self.controlled_vehicles[vehicle].publish_cmd(v, omega)


###########################################
########### Central Controllers ###########
###########################################

# In run_live.py these are called once per update step.

class intersection_demo():
    # self, controller_parameters: dict, optitrack_parameters: dict, path_generator, controlled_vehicles, goal_paths, delta_t
    def __init__(self, controller_parameters: dict, path_parameters: dict, path_generator, controlled_vehicles, goal_paths, delta_t,obstacles:dict,fleet_info:dict):
        
        self.v_bar:float        = controller_parameters["v_bar"]
        self.distance2front     = controller_parameters['distance2front']
        
        self.buffer_points      = int(path_parameters['path_points_per_m'] * controller_parameters['path_buffer'])
        self.lookahead_points   = int(path_parameters['path_points_per_m'] * controller_parameters['lookahead_distance'])

        # Take time step to be constant during experiment
        self.delta_t = delta_t

        self.goal_paths             = goal_paths
        self.intersection_waypoints = path_generator.intersection_waypoints

        self.controlled_vehicles = controlled_vehicles

        self.fleet_info = fleet_info
        self.fleets = {
            fleet: {'leader': members[0], 'followers': members[1:] if len(members) > 1 else []} 
            for fleet, members in fleet_info.items() if members
        }


        self.fleet_roles={}
        for fleet,members in self.fleets.items():
            self.fleet_roles[members['leader']]='leader'
            for follower in members['followers']:
                self.fleet_roles[follower]='follower'

        self.vehicles_control_info = {
            vehicle: {"closest_path_point_index": None, "target_pos": None,
            "at_intersection": False, "intersection_waypoint": None,"obstacle_detected": False,"State":"normal",
            "lateral_controller": PID(controller_parameters['PID']),
            "total_time_in_obstacle":0,
            "total_distance_in_obstacle":0,
            "last_position":0,
            "last_state":None,
            "last_timestamp":0,
            "past_positions":[]
            }
            for vehicle in self.controlled_vehicles.keys()
        }
        
        # Central controller variables
        self.intersection_queue = []
        self.intersection_timeoff = controller_parameters['intersection_timeoff']
        self.intersection_timer_start = time.time()
        self.intersection_timer_ellapsed = None
        self.intersection_stoptime = {}

        self.obstacles = obstacles
        self.vehicle_positions={vehicle: [] for vehicle in self.controlled_vehicles}
        self.iteration_count=0
        self.legend = False
        self.figures={}
        self.start_time=datetime.datetime.now()
        

    def plot_path(self, vehicle):
        vehicle_positions = np.array(self.vehicle_positions[vehicle])
        target_pos = np.array(self.vehicles_control_info[vehicle]['target_pos'])
        
        plt.plot(target_pos[0], target_pos[1], 'go', label='target path')
        plt.plot(vehicle_positions[:, 0], vehicle_positions[:, 1], 'b-', label='Vehicle Path')

        for obstacle_key, obstacle_position in self.obstacles.items():
            plt.plot(obstacle_position[0], -obstacle_position[1], 'ro', label='Obstacle')
        
        if hasattr(self,'rrt_path'):
            plt.plot(self.rrt[:,0],self.rrt[:,1],'k',label='RRT')

        plt.title(f'Path of Vehicle {vehicle}')
        plt.xlabel('X')
        plt.ylabel('Y')
        plt.xlim(0,3)
        plt.ylim(-3,0)

        if not self.legend:
            plt.legend()
            self.legend = True

        plt.pause(0.1)
        plt.show(block=False) 

    def control_action(self, vehicle, distance2front_vehicle):
        """ Uses info/values obtained from the control_step() to calculate the
            desired linear and angular velocities for the vehicles.
        """
        vehicle_key=f'td{str(vehicle).zfill(2)}'
        jerk=0
        TTC=0
        v=0
        omega=0           
        if self.fleet_roles[vehicle_key]=='leader': #leader

            # Base value for v
            v = self.v_bar

            # TODO: why does self.controlled_vehicles[vehicle].position not work? need self.controlled_vechiles[vehicle].position?
            angle2target = dyn.angle(self.vehicles_control_info[vehicle]['target_pos'] - self.controlled_vehicles[vehicle].position, radians=True)
            heading_error = angle_within_range(angle2target - self.controlled_vehicles[vehicle].angle) # Range [-180, +180]
            
            # Base value for omega
            omega = self.vehicles_control_info[vehicle]['lateral_controller'].control_step(heading_error, self.delta_t) # PID controller

            state = self.vehicles_control_info[vehicle]['state']

            # If too close to front vehicle
            if (distance2front_vehicle < self.distance2front):
                v = 0.0
                omega = 0.0


            if state == 'intersection':
            # If vehicle at intersection
                if (self.vehicles_control_info[vehicle]['at_intersection']):
                    print(f"vehicle {vehicle} at intersection.")
                    distance2_intersect = np.linalg.norm(self.vehicles_control_info[vehicle]['intersection_waypoint'] - self.controlled_vehicles[vehicle].position)

                    # If not close enough to stop line slow down to reach it.
                    if (distance2_intersect > 0.10):
                        v = self.v_bar * 0.9
                        print("slowing down pre_intersection")
                    else:
                        print("stopping at intersection")
                        v = 0.0
                        omega = 0.0       


            elif state == 'obstacle':
            # If too close to obstacle
                omega_default = omega
                v_default = v
                cond = False

                for obstacle_key, obstacle in self.obstacles.items():
                    dx, dy  = abs(self.controlled_vehicles[vehicle].position[0])-obstacle[0],abs(self.controlled_vehicles[vehicle].position[1])-obstacle[1]
                    distance = np.sqrt(dx**2 + dy**2)
                    print(self.controlled_vehicles[vehicle].position)
                    print(distance)
                    if distance <= 0.5:
                        # RRT

                        cond = True
                        print('obstacle detected')
                        rrt = RRT()
                        path = rrt.rrt(self.controlled_vehicles[vehicle].position,self.vehicles_control_info[vehicle]['target_pos'],obstacle)

                        dy_target = path[1][1]- self.controlled_vehicles[vehicle].position[1] 
                        dx_target = path[1][0]- self.controlled_vehicles[vehicle].position[0]
                                        
                        angle_target = math.atan2(dy_target, dx_target)
                        print('this is angle target',angle_target)
                        omega = angle_target
                        v = 0.3

                        # # Circular

                        # cond = True
                        # radius = 0.75
                        # print ('obstacle detected')
                        # arc = distance*math.pi
                        # omega = -arc/2               
                        # v=0.3    
                        # # v = max(abs(omega*radius),0.3)
                        # break

                    else:
                        cond = False

                    if not cond:
                        print('no more obstacle')
                        # omega = omega_default
                        # v = v_default
                    else:
                        dy_target = self.vehicles_control_info[vehicle]['target_pos'][1] - self.controlled_vehicles[vehicle].position[1]
                        dx_target = self.vehicles_control_info[vehicle]['target_pos'][0] - self.controlled_vehicles[vehicle].position[0]
                        angle_target = math.atan2(dy_target, dx_target)
                        omega = angle_target

                # Track time and distance
                current_time=time.time()
                elapsed_time = current_time - self.vehicles_control_info[vehicle]['last_timestamp']
                current_position = self.controlled_vehicles[vehicle].position
                last_position = self.vehicles_control_info[vehicle]['last_position']
                distance_travelled = np.linalg.norm(np.array(current_position) - np.array(last_position))
                if self.vehicles_control_info[vehicle]['last_state'] == 'obstacle':
                    self.vehicles_control_info[vehicle]['total_time_in_obstacle'] += elapsed_time
                    self.vehicles_control_info[vehicle]['total_distance_in_obstacle'] += distance_travelled
                self.vehicles_control_info[vehicle]['last_timestamp'] = current_time
                self.vehicles_control_info[vehicle]['last_position'] = current_position


            else: #Normal mode
                
                v = self.v_bar
                omega = self.vehicles_control_info[vehicle]['lateral_controller'].control_step(heading_error, self.delta_t)       

            # write into txt file to determine time and distance in obstacle state
            self.vehicles_control_info[vehicle]['last_state'] = state
            with open('vehicle_info.txt', 'a') as file:
                file.write(f"Vehicle {vehicle}\n")
                file.write(f"Total time in 'obstacle' state: {self.vehicles_control_info[vehicle]['total_time_in_obstacle']} seconds\n")
                file.write(f"Distance travelled in 'obstacle' state: {self.vehicles_control_info[vehicle]['total_distance_in_obstacle']} units\n") 

            # # write into excel file for plotting
            # filename='vehicle_info.xlsx'

            # try:
            #     wb=load_workbook(filename)
            # except FileNotFoundError:
            #     wb=openpyxl.Workbook()
            
            # for vehicle_id in self.controlled_vehicles:
            #     vehicle_key=f'td{str(vehicle).zfill(2)}'

            #     sheet = wb[vehicle_key] if vehicle_key in wb.sheetnames else wb.create_sheet(title=vehicle_key)
                
            #     row=sheet.max_row+1

            #     current_location = self.controlled_vehicles[vehicle].position
            #     target_location = self.vehicles_control_info[vehicle]['target_pos']
            #     sheet[f"A{row}"] = current_location[0]  
            #     sheet[f"B{row}"] = current_location[1]
            #     sheet[f"C{row}"] = target_location[0]  
            #     sheet[f"D{row}"] = target_location[1] 

            #     sheet["A1"]="Current Location X"
            #     sheet["B1"]="Current Location Y"
            #     sheet["C1"]="Target Location X"
            #     sheet["D1"]="Target Location Y"     

            #     sheet.column_dimensions['A'].width=20                          
            #     sheet.column_dimensions['B'].width=20
            #     sheet.column_dimensions['C'].width=20                
            #     sheet.column_dimensions['D'].width=20  

            # obstacle_sheet = wb["Obstacles"] if "Obstacles" in wb.sheetnames else wb.create_sheet(title="Obstacles")
            # if obstacle_sheet.max_row==1:
            #     obstacle_sheet["A1"] = "Obstacle"
            #     obstacle_sheet["B1"] = "Location X"
            #     obstacle_sheet["C1"] = "Location Y"

            #     obstacle_sheet.column_dimensions['A'].width=20                          
            #     obstacle_sheet.column_dimensions['B'].width=20
            #     obstacle_sheet.column_dimensions['C'].width=20  

            #     row = obstacle_sheet.max_row + 1
            #     for obstacle_key, obstacle_location in self.obstacles.items():
            #         obstacle_sheet[f"A{row}"] = obstacle_key
            #         obstacle_sheet[f"B{row}"] = obstacle_location[0]
            #         obstacle_sheet[f"C{row}"] = obstacle_location[1]           
            #         row += 1    
 
            # wb.save(filename)

        elif self.fleet_roles[vehicle_key]=='follower': #follower
            leader_vehicle=None
            leader_pos=None
            leader_angle=None

            for fleet, members in self.fleet_info.items():
                if vehicle_key in members:
                    leader_vehicle = int(members[members.index(vehicle_key)-1][2:])
                if leader_vehicle in self.controlled_vehicles:
                    leader = self.controlled_vehicles[leader_vehicle]
                    print(leader)
                    leader_pos = leader.position
                    leader_angle = leader.angle
                    
                    # # Store the current position of the vehicle on each timestep
                    # past_positions = self.vehicles_control_info[vehicle]['past_positions']
                    # past_positions.append(self.controlled_vehicles[vehicle].position)

                    # delay = 50

            
                    # if len(past_positions) > delay:
                    #     leader_pos = self.vehicles_control_info[leader_vehicle]['past_positions'][0]
                    #     self.vehicles_control_info[leader_vehicle]['past_positions'].pop(0)


                    break
    
            print(f"Leader of fleet {fleet}: {leader_vehicle}")
            print(f"Leader position: {leader_pos}")
            print(f"Leader angle: {leader_angle}")

            if leader_pos is not None and leader_angle is not None:
                vehicle_key=int(vehicle_key[2:])
                # Base value for v
                v = self.v_bar


                desired_distance = 0.5 


                dy = leader_pos[1] - self.controlled_vehicles[vehicle_key].position[1]
                dx = leader_pos[0] - self.controlled_vehicles[vehicle_key].position[0]
                angle_to_leader = math.atan2(dy, dx)


                distance_to_leader = np.linalg.norm(np.array(self.controlled_vehicles[vehicle_key].position) - np.array(leader_pos))

    
                if distance_to_leader > desired_distance:
                    v = self.v_bar * 1.05
                    v= max(v,0.6)
                elif distance_to_leader < desired_distance:
                    v = self.v_bar * 0.1 
                elif distance_to_leader < 0.3:
                    v = 0


                angle_error = angle_within_range(angle_to_leader - self.controlled_vehicles[vehicle_key].angle)
            
                omega = 0.3*self.vehicles_control_info[vehicle]['lateral_controller'].control_step(angle_error, self.delta_t)                    


                vehicle_info=self.vehicles_control_info[vehicle_key]

                vel_diff=v-vehicle_info.get('previous_velocity',v)
                current_a=vel_diff/self.delta_t

                acc_diff=current_a - vehicle_info.get('previous_acceleration',current_a)
                jerk=acc_diff/self.delta_t

                vehicle_info['previous_velocity']=v
                vehicle_info['previous_acceleration']=current_a

                TTC=100
                if v!=0:
                    TTC=distance_to_leader/v           



        else:
            print('There has been an Error')
            v=0
            omega=0


        # write into excel file for plotting
        filename='vehicle_info.xlsx'

        try:
            wb=load_workbook(filename)
        except FileNotFoundError:
            wb=openpyxl.Workbook()
        
        for vehicle_id in self.controlled_vehicles:
            vehicle_key=f'td{str(vehicle_id).zfill(2)}'

            sheet = wb[vehicle_key] if vehicle_key in wb.sheetnames else wb.create_sheet(title=vehicle_key)
            
            row=sheet.max_row+1

            current_time=datetime.datetime.now()
            elapsed_time=(current_time-self.start_time).total_seconds()


            current_location = self.controlled_vehicles[vehicle].position
            target_location = self.vehicles_control_info[vehicle]['target_pos']
            sheet[f"A{row}"] = current_location[0]  
            sheet[f"B{row}"] = current_location[1]
            sheet[f"C{row}"] = target_location[0]  
            sheet[f"D{row}"] = target_location[1] 

            if self.fleet_roles[vehicle_key]!='leader':
                sheet[f"E{row}"] = jerk
                sheet[f"F{row}"] = TTC
            sheet[f"G{row}"]=elapsed_time

            sheet["A1"]="Current Location X"
            sheet["B1"]="Current Location Y"
            sheet["C1"]="Target Location X"
            sheet["D1"]="Target Location Y"     
            if self.fleet_roles[vehicle_key]!='leader':
                sheet["E1"] = "da/dt"
                sheet["F1"] = "TTC"     
            sheet["G1"]="Time"

            sheet.column_dimensions['A'].width=20                          
            sheet.column_dimensions['B'].width=20
            sheet.column_dimensions['C'].width=20                
            sheet.column_dimensions['D'].width=20
            sheet.column_dimensions['E'].width=20                
            sheet.column_dimensions['F'].width=20    
            sheet.column_dimensions['G'].width=20   

        obstacle_sheet = wb["Obstacles"] if "Obstacles" in wb.sheetnames else wb.create_sheet(title="Obstacles")
        if obstacle_sheet.max_row==1:
            obstacle_sheet["A1"] = "Obstacle"
            obstacle_sheet["B1"] = "Location X"
            obstacle_sheet["C1"] = "Location Y"

            obstacle_sheet.column_dimensions['A'].width=20                          
            obstacle_sheet.column_dimensions['B'].width=20
            obstacle_sheet.column_dimensions['C'].width=20  

            row = obstacle_sheet.max_row + 1
            for obstacle_key, obstacle_location in self.obstacles.items():
                obstacle_sheet[f"A{row}"] = obstacle_key
                obstacle_sheet[f"B{row}"] = obstacle_location[0]
                obstacle_sheet[f"C{row}"] = obstacle_location[1]           
                row += 1    

        wb.save(filename)


            
        return v, omega


    def control_step(self):
        for vehicle in self.controlled_vehicles:

            self.vehicles_control_info[vehicle]['past_positions']=[]
            # Selecting sub-section of path where the goal point can be located.
            ## Because of path overlap at intersection, basic min distance would not work.
            if self.vehicles_control_info[vehicle]['closest_path_point_index'] is None:
                goal_path_subset = self.goal_paths[vehicle]

            else:
                path_buffer_end_idx = self.vehicles_control_info[vehicle]['closest_path_point_index'] + self.buffer_points

                # TODO: add something for when path_loop = False
                if path_buffer_end_idx < len(self.goal_paths[vehicle]):
                    goal_path_subset = self.goal_paths[vehicle][self.vehicles_control_info[vehicle]['closest_path_point_index']:path_buffer_end_idx]
                else:
                    overflow = path_buffer_end_idx - len(self.goal_paths[vehicle])
                    goal_path_subset = self.goal_paths[vehicle][self.vehicles_control_info[vehicle]['closest_path_point_index']:len(self.goal_paths[vehicle])]
                    goal_path_subset = np.concatenate([goal_path_subset, self.goal_paths[vehicle][:overflow]])
            ###############


            # Logic for permission to go on intersection
            self.intersection_timer_ellapsed = time.time() - self.intersection_timer_start
            if (len(self.intersection_queue) > 0):
                if (self.intersection_timer_ellapsed > self.intersection_timeoff):
                    crossing_vehicle = self.intersection_queue.pop(0) # 1st vehicle in list was the 1st at the intersection
                    self.vehicles_control_info[crossing_vehicle]['at_intersection'] = False

                    self.intersection_timer_ellapsed = 0.0
                    self.intersection_timer_start = time.time()

                    print(f"{crossing_vehicle} not at intersection anymore")
                else:
                    print(f"{self.intersection_queue} still waiting at intersection.")
            


            # Getting index of closest point in sub-path
            distance_to_subpath_points = np.linalg.norm(self.controlled_vehicles[vehicle].position - goal_path_subset, axis=1)
            min_distance = np.min(distance_to_subpath_points)
            closest_pathpoint_idx = np.where(distance_to_subpath_points == min_distance)[0][0]

            # Getting index of closest point in goal path
            if self.vehicles_control_info[vehicle]['closest_path_point_index'] is not None:
                closest_pathpoint_idx = (closest_pathpoint_idx + self.vehicles_control_info[vehicle]['closest_path_point_index']) % (len(self.goal_paths[vehicle])-1)

            self.vehicles_control_info[vehicle]['closest_path_point_index'] = closest_pathpoint_idx


            # Getting target position
            target_idx = (self.vehicles_control_info[vehicle]['closest_path_point_index'] + self.lookahead_points) % (len(self.goal_paths[vehicle]) - 1)
            self.vehicles_control_info[vehicle]['target_pos'] = self.goal_paths[vehicle][target_idx]


            # Checking if car at intersection
            # TODO: there is for sure a better way to do this (compare floats)
            approx_target_pos = [np.round(elem,2) for elem in self.vehicles_control_info[vehicle]['target_pos']] 
            #print(approx_target_pos)
            #print(np.round(self.intersection_waypoints,2))

            for intersection_pos in np.round(self.intersection_waypoints,2):
                if (approx_target_pos == intersection_pos).all():
                    print(f"reached intersection, approx_target={approx_target_pos}, intersec = {np.round(self.intersection_waypoints,2)}")

                    self.vehicles_control_info[vehicle]['intersection_waypoint'] = self.vehicles_control_info[vehicle]['target_pos']
                    self.vehicles_control_info[vehicle]['at_intersection'] = True
                    self.intersection_queue.append(vehicle)

            obstacle_detected = False
            for obstacle_key,obstacle_position in self.obstacles.items():
                if np.linalg.norm(abs(self.controlled_vehicles[vehicle].position) - obstacle_position)<0.5:
                    obstacle_detected = True
                    break
            self.vehicles_control_info[vehicle]['obstacle_detection']=obstacle_detected

            if self.vehicles_control_info[vehicle]['at_intersection']:
                self.vehicles_control_info[vehicle]['state'] = 'intersection'            
            elif self.vehicles_control_info[vehicle]['obstacle_detection']:
                self.vehicles_control_info[vehicle]['state'] = 'obstacle'
            else:
                self.vehicles_control_info[vehicle]['state'] = 'normal'               
            print(self.vehicles_control_info[vehicle]['state'])

            # if approx_target_pos in np.round(self.intersection_waypoints,2):
            #     if not self.vehicles_control_info[vehicle]['at_intersection']:

            #         print(f"reached intersection, approx_target={approx_target_pos}, intersec = {np.round(self.intersection_waypoints,2)}")
            #         self.vehicles_control_info[vehicle]['intersection_waypoint'] = self.vehicles_control_info[vehicle]['target_pos']
            #         self.vehicles_control_info[vehicle]['at_intersection'] = True
            #         self.intersection_queue.append(vehicle)


            # Getting distance to car in front
            if all(self.vehicles_control_info[vehicle]['closest_path_point_index'] is not None for vehicle in self.vehicles_control_info):
                other_vehicles_pos = [self.controlled_vehicles[other_vehicle].position for other_vehicle in self.controlled_vehicles
                                    if (self.controlled_vehicles[other_vehicle] != self.controlled_vehicles[vehicle])
                                    and (self.goal_paths[other_vehicle].shape == self.goal_paths[vehicle].shape)
                                    and (self.vehicles_control_info[other_vehicle]['closest_path_point_index'] > self.vehicles_control_info[vehicle]['closest_path_point_index'])]  #TODO: this is broken because of looping over index

                # TODO: why is this exception here?
                # TODO: change all np.subtract to np.array subtractions
                try: distance2otherVehicles = np.linalg.norm(np.subtract(other_vehicles_pos, self.controlled_vehicles[vehicle].position), axis=1)
                except: distance2otherVehicles = np.array([])

                if len(distance2otherVehicles >=1): distance2front_vehicle = np.min(distance2otherVehicles)
                else: distance2front_vehicle = 999999

            else: distance2front_vehicle = 99999999

            # Using info obtained above to get control action
            v, omega = self.control_action(vehicle, distance2front_vehicle)
            # Robot object converts (v,omega) to the required ROS command
            self.controlled_vehicles[vehicle].publish_cmd(v, omega)
            
            print(f"v: {v}, omega: {omega}")
            

            # Store vehicle positions
            self.vehicle_positions[vehicle].append(self.controlled_vehicles[vehicle].position)
            
            # # Update the plot
            # if self.iteration_count % 10 == 0:
            #     self.plot_path(vehicle)
            # self.iteration_count += 1

########################
### Helper Functions ####
#########################


def angle_within_range(angle, rad=True, min_val=-np.pi, max_val=+np.pi):
    while angle < min_val:
        if rad: angle += 2.0*np.pi
        else:   angle += 360.0
    
    while angle > max_val:
        if rad: angle -= 2.0*np.pi
        else:   angle -= 360.0

    return angle