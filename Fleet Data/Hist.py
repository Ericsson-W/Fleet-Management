import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
import numpy 
import scipy.stats as stats

wb='Test2.xlsx'
sheetnames = load_workbook(wb).sheetnames


for i in range(3,len(sheetnames)-1):            
    data = pd.read_excel(wb,sheet_name=i,engine='openpyxl')
    sheetname=sheetnames[i]


    kurtosis_value = stats.kurtosis(data['da/dt'])
    print(f'the kurtosis value of jerk of {sheetname} is',kurtosis_value)
    fig, axs = plt.subplots(2)

    axs[0].hist(data['da/dt'], bins=30, alpha=0.5)
    axs[0].set_title(f'Distribution of Jerk of {sheetname}')
    axs[0].set_xlabel('Jerk[$\mathregular{ms^{-3}}$]')

    axs[1].hist(data['TTC'], bins=30, alpha=0.5)
    axs[1].set_title(f'Distribution of TTC of {sheetname}')
    axs[1].set_xlabel('time[s]')

    mean_TTC = numpy.mean(data['TTC'])
    print(f'the mean TTC value of {sheetname} is',mean_TTC,'seconds')

    LQR=numpy.percentile(data['TTC'],25)
    UQR=numpy.percentile(data['TTC'],75)
    print(f'the 25th percentile of TTC of {sheetname} is',LQR)
    print(f'the 75th percentile of TTC of {sheetname} is',UQR)

    plt.tight_layout()  
    plt.show() 
    plt.pause(0.1)

    fig, axs = plt.subplots(2)

    axs[0].plot(data['Time'],data['da/dt'],'b')
    axs[0].set_title(f'Jerk of {sheetname} against time')
    axs[0].set_ylabel('Jerk[$\mathregular{s^{-3}}$]')
    axs[0].set_xlabel('time[s]')
    axs[0].legend()


    axs[1].plot(data['Time'],data['TTC'],'b')
    axs[1].set_title(f'TTC of {sheetname} against time')
    axs[1].set_ylabel('TTC[s]')
    axs[1].set_xlabel('time[s]')
    axs[1].legend()

    fig.tight_layout()

    plt.show()